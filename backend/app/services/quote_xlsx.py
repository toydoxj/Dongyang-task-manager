"""견적서 xlsx 출력 — 사장 운영 양식(`templates/quote_template.xlsx`)에 셀 fill.

사용자가 입력한 값(form_data['input'])을 template의 알려진 셀 좌표에 채워 넣고,
산출 셀(K18~K30 등)의 Excel 수식은 그대로 두어 Excel이 열릴 때 자동 계산.
산출 결과는 quote_form_data['result']에도 저장되지만 xlsx에는 수식 유지가 정답
(원본과 동일한 인쇄·재계산 동작).

template 셀 좌표는 docs/설계견적26-01-007*.xlsx 셀 dump 분석 결과:
  B3  문서번호 (예: " 26 - 01 - 007")
  D4  수신처 회사명          J4  전화
  D5  참조자                 J5  E-mail
  D6  용역명
  D7  위치
  E8  연면적 (number)
  J8  층수 텍스트
  D9  구조형식
  J12 계수 (0.5/1.0)
  H16 종별 요율
  H17 구조방식 요율
  K21 보고서인쇄비           K22 추가조사비
  I23 교통비 인.일
  H28 당사조정 %
  D31 지불방법
  D32 특이사항
  G33 작성일
"""
from __future__ import annotations

import io
from copy import copy
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.quote_calculator import _excel_round_half_up

_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "quote_template.xlsx"
_SHEET_NAME = "구조검토"


def _insert_rows_preserving_merges(
    ws: Worksheet, idx: int, amount: int
) -> None:
    """openpyxl insert_rows의 한계(merge cells 미이동) 보강.

    idx 이상에서 시작하는 merge ranges를 amount만큼 row down shift.
    cell value/style은 openpyxl이 자동 이동시키지만 merged_cells.ranges는
    내부 row index를 갱신하지 않아 새로 unmerge → insert → re-merge 필요.

    교차 merge(min_row < idx <= max_row)는 분할 정책이 양식별로 다르므로
    감지 시 즉시 예외 — 호출자가 idx를 조정해야 함.
    """
    affected: list[str] = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row < idx <= mr.max_row:
            raise ValueError(
                f"insert idx={idx}가 기존 merge {mr} 안에 걸쳐 있습니다 — "
                "분할 정책 미구현. 다른 idx로 호출하거나 helper를 확장하세요."
            )
        if mr.min_row >= idx:
            min_col = get_column_letter(mr.min_col)
            max_col = get_column_letter(mr.max_col)
            shifted = (
                f"{min_col}{mr.min_row + amount}:"
                f"{max_col}{mr.max_row + amount}"
            )
            ws.unmerge_cells(str(mr))
            affected.append(shifted)
    ws.insert_rows(idx, amount=amount)
    for r in affected:
        ws.merge_cells(r)


def _copy_row_style(
    ws: Worksheet, src_row: int, dst_rows: list[int], cols: str
) -> None:
    """src_row의 cell style을 dst_rows 각 row에 복사 (border/font/fill/format)."""
    for col in cols:
        src = ws[f"{col}{src_row}"]
        for dst_row in dst_rows:
            ws[f"{col}{dst_row}"]._style = copy(src._style)


def build_quote_xlsx(
    form_data: dict[str, Any], *, doc_number: str = ""
) -> bytes:
    """form_data['input']를 template에 fill 후 bytes 반환.

    Excel 수식(K18~K30, D10 NUMBERSTRING 등)은 그대로 보존되어 Excel/Numbers/
    LibreOffice가 열 때 자동 재계산된다.
    """
    if not _TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"견적서 템플릿이 없습니다: {_TEMPLATE_PATH}. "
            "backend/app/templates/quote_template.xlsx 를 배포에 포함하세요."
        )

    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb[_SHEET_NAME]
    inp = form_data.get("input") or {}

    # 헤더
    if doc_number:
        # 사장 양식의 공백 패턴(" 26 - 01 - 007") 유지
        yy, mm, nnn = doc_number.split("-")
        ws["B3"] = f" {yy} - {mm} - {nnn}"

    # 수신처
    ws["D4"] = inp.get("recipient_company", "")
    ws["D5"] = inp.get("recipient_person", "")
    if inp.get("recipient_phone"):
        ws["J4"] = inp["recipient_phone"]
    if inp.get("recipient_email"):
        ws["J5"] = inp["recipient_email"]

    # 용역 정보
    ws["D6"] = inp.get("service_name", "")
    ws["D7"] = inp.get("location", "")
    if inp.get("gross_floor_area") is not None:
        ws["E8"] = float(inp["gross_floor_area"])
    if inp.get("floors_text"):
        ws["J8"] = inp["floors_text"]
    if inp.get("structure_form"):
        ws["D9"] = f"  {inp['structure_form']}"  # template은 두 칸 들여쓴 형태

    # 산출 변수
    ws["J12"] = float(inp.get("coefficient", 1.0))
    ws["H16"] = float(inp.get("type_rate", 1.0))
    ws["H17"] = float(inp.get("structure_rate", 1.0))
    ws["H28"] = int(inp.get("adjustment_pct", 87))

    # 직접경비
    ws["K21"] = int(inp.get("printing_fee") or 0)
    ws["K22"] = int(inp.get("survey_fee") or 0)
    ws["I23"] = int(inp.get("transport_persons") or 0)

    # VAT 포함 표시 — K30(용역대가) 다음에 3행 삽입(공급가액/VAT/합계).
    # row 30 layout: B30:C30(번호) D30:J30(라벨 merge) K30:L30(값 merge).
    # 새 row 31~33도 동일 layout으로 merge + 스타일 복사하여 표 디자인 보존.
    # PDF 전환 시 이 분기 재작성 예정.
    vat_included = bool(inp.get("vat_included"))
    if vat_included:
        result = form_data.get("result") or {}
        final_amount = int(result.get("final") or 0)
        vat_amt = int(
            result.get("vat_amount")
            or _excel_round_half_up(final_amount * 0.1, 0)
        )
        final_wv = int(
            result.get("final_with_vat") or final_amount + vat_amt
        )
        _insert_rows_preserving_merges(ws, 31, 3)
        for r in (31, 32, 33):
            ws.merge_cells(f"D{r}:J{r}")
            ws.merge_cells(f"K{r}:L{r}")
        _copy_row_style(ws, 30, [31, 32, 33], "BCDEFGHIJKL")
        # row height 보존 — _style은 셀 단위만 복사하므로 row dimension은 별도
        src_h = ws.row_dimensions[30].height
        if src_h:
            for r in (31, 32, 33):
                ws.row_dimensions[r].height = src_h
        ws["D31"] = "공 급 가 액 (VAT 별도)"
        ws["K31"] = final_amount
        ws["D32"] = "V A T  (10%)"
        ws["K32"] = vat_amt
        ws["D33"] = "합     계  (VAT 포함)"
        ws["K33"] = final_wv
        # print_area 확장 — 기존 A1:L39 → A1:L42 (3행 추가)
        if ws.print_area:
            ws.print_area = "A1:L42"

    # 자유 텍스트 — vat 행 삽입한 경우 D31/D32가 D34/D35로 밀림
    terms_row = 34 if vat_included else 31
    notes_row = 35 if vat_included else 32
    date_row = 36 if vat_included else 33

    if inp.get("payment_terms"):
        ws[f"D{terms_row}"] = f" {inp['payment_terms']}"
    if inp.get("special_notes"):
        ws[f"D{notes_row}"] = f" {inp['special_notes']}"

    # 작성일 (오늘 KST)
    ws[f"G{date_row}"] = "            " + date.today().strftime("%Y. %m. %d")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def quote_filename(doc_number: str, service_name: str) -> str:
    """파일명 규칙: 설계견적{doc_number}({service_name 안전한 일부}).xlsx.

    윈도우 금지 문자(\\/:*?"<>|) + 줄바꿈 제거. 30자 cap.
    """
    safe = "".join(c for c in service_name if c not in r'\/:*?"<>|' + "\r\n")
    safe = safe.strip()[:30]
    return f"설계견적{doc_number}({safe}).xlsx"
