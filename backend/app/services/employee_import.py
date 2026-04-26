"""엑셀 → employees 화이트리스트 import.

민감 컬럼(연봉/실적/인상률 등)은 일체 읽지 않는다 — 헤더 매칭으로 안전 컬럼만 추출.
파일은 메모리에서만 처리, 디스크에 저장 X.
"""
from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import load_workbook

# 안전한 컬럼만 — 엑셀 헤더의 변형(공백/숫자 접미)을 정규화 후 매칭
# value: 우리 모델 필드명
SAFE_COLUMN_MAP: dict[str, str] = {
    "이름": "name",
    "성명": "name",
    "직급": "position",  # "2026 직급2" 같은 변형도 _normalize로 매칭
    "소속": "team",
    "팀": "team",
    "부서": "team",
    "학위": "degree",
    "자격": "license",
    "면허": "license",
    "등급": "grade",
    "이메일": "email",
    "email": "email",
    "e-mail": "email",
}


@dataclass
class ParsedEmployee:
    name: str
    position: str = ""
    team: str = ""
    degree: str = ""
    license: str = ""
    grade: str = ""
    email: str = ""


def _normalize(s: object) -> str:
    if s is None:
        return ""
    t = str(s).strip().lower()
    # 공백 제거 + 숫자 접미 제거 (헤더 자동 번호 ex. "직급2")
    t = "".join(ch for ch in t if not ch.isspace())
    while t and t[-1].isdigit():
        t = t[:-1]
    return t


def _build_normalized_map() -> dict[str, str]:
    return {_normalize(k): v for k, v in SAFE_COLUMN_MAP.items()}


def parse_workbook(content: bytes) -> list[ParsedEmployee]:
    """엑셀 첫 시트 파싱. 헤더 행을 자동 탐지(이름/성명 컬럼이 있는 행)."""
    wb = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    norm_map = _build_normalized_map()

    # 헤더 행 자동 탐지: 첫 20행 중 "이름" 또는 "성명"이 있는 행
    header_row_idx: int | None = None
    header_cells: list[str] = []
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        normed = [_normalize(c) for c in row]
        if "이름" in normed or "성명" in normed:
            header_row_idx = ridx
            header_cells = [str(c) if c is not None else "" for c in row]
            break

    if header_row_idx is None:
        return []

    # 헤더 → (column_index, our_field) 매핑 (안전 컬럼만)
    safe_cols: list[tuple[int, str]] = []
    for cidx, raw in enumerate(header_cells):
        field = norm_map.get(_normalize(raw))
        if field:
            safe_cols.append((cidx, field))

    if not safe_cols:
        return []

    name_field_indices = [ci for ci, f in safe_cols if f == "name"]
    if not name_field_indices:
        return []

    results: list[ParsedEmployee] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # 화이트리스트 컬럼만 dict에 모음
        rec: dict[str, str] = {}
        for cidx, field in safe_cols:
            if cidx >= len(row):
                continue
            v = row[cidx]
            if v is None:
                continue
            rec[field] = str(v).strip()

        name = rec.get("name", "").strip()
        if not name:
            continue
        results.append(
            ParsedEmployee(
                name=name,
                position=rec.get("position", ""),
                team=rec.get("team", ""),
                degree=rec.get("degree", ""),
                license=rec.get("license", ""),
                grade=rec.get("grade", ""),
                email=rec.get("email", ""),
            )
        )
    return results
