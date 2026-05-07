"""견적서 xlsx 셀 dump 헬퍼 (PR-Q0).

운영자가 종류별 1회 실행 → 산출식·요율·드롭다운 옵션 markdown 추출 →
`docs/quote_formulas/{type}.md` 보관 → PR-Q2~Q9에서 strategy 함수에 transcribe.

Usage:
    uv run python scripts/quote_xlsx_dump.py <xlsx_path> [output.md]

예:
    uv run python scripts/quote_xlsx_dump.py ../docs/구조감리26-08-003*.xlsx \\
        ../docs/quote_formulas/supervision.md

.xls 파일(.xls ≠ .xlsx)은 openpyxl이 못 읽으므로 LibreOffice로 변환:
    soffice --headless --convert-to xlsx '../docs/정기점검견적26-02-23.xls'

검증 메모:
    formula 컬럼이 모두 비어있으면 산출 결과만 보존된 사본일 가능성 →
    스크립트가 alert을 출력. 그 경우 원본 sheet 별도 확보 필요.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def _truncate(text: str, limit: int) -> str:
    """markdown table 셀에 안전하게 들어가도록 줄바꿈/파이프 escape + 길이 cap."""
    if text is None:
        return ""
    s = str(text).replace("\n", " ⏎ ").replace("|", "\\|")
    if len(s) > limit:
        s = s[:limit] + "…"
    return s


def dump(xlsx_path: Path, out: Path | None = None) -> str:
    """xlsx의 모든 sheet를 markdown으로 dump."""
    # data_only=False: 수식 그대로. data_only=True: 마지막 저장 시점 값.
    wb_formula = load_workbook(xlsx_path, data_only=False)
    wb_value = load_workbook(xlsx_path, data_only=True)

    lines: list[str] = []
    lines.append(f"# {xlsx_path.name}")
    lines.append("")
    lines.append(f"- 경로: `{xlsx_path}`")
    lines.append(f"- Sheet 수: {len(wb_formula.sheetnames)}")
    lines.append("")

    for sheet_name in wb_formula.sheetnames:
        ws_f = wb_formula[sheet_name]
        ws_v = wb_value[sheet_name]
        lines.append(
            f"## Sheet: `{sheet_name}` "
            f"({ws_f.max_row} rows × {ws_f.max_column} cols)"
        )
        lines.append("")

        # 1. 병합 셀
        merges = sorted(
            ws_f.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)
        )
        if merges:
            lines.append("### Merged Cells")
            lines.append("")
            for mr in merges:
                lines.append(f"- `{mr}`")
            lines.append("")

        # 2. 데이터 검증 (드롭다운 옵션)
        dvs = list(ws_f.data_validations.dataValidation)
        if dvs:
            lines.append("### Data Validations (드롭다운 옵션 등)")
            lines.append("")
            lines.append("| 영역 | type | formula1 | formula2 |")
            lines.append("|---|---|---|---|")
            for dv in dvs:
                lines.append(
                    f"| {dv.sqref} | {dv.type or ''} | "
                    f"{_truncate(dv.formula1, 80)} | "
                    f"{_truncate(dv.formula2, 40)} |"
                )
            lines.append("")

        # 3. 셀 dump — 값 또는 수식이 있는 셀만
        formula_cells: list[tuple[str, str, str]] = []
        value_cells: list[tuple[str, str]] = []
        for row in ws_f.iter_rows():
            for cell_f in row:
                raw = cell_f.value
                if raw is None or (isinstance(raw, str) and raw == ""):
                    continue
                coord = cell_f.coordinate
                # data_only=True 워크북에서 동일 좌표의 계산값
                cell_v = ws_v[coord]
                computed = cell_v.value
                if isinstance(raw, str) and raw.startswith("="):
                    formula_cells.append(
                        (coord, _truncate(raw, 100), _truncate(computed, 60))
                    )
                else:
                    value_cells.append((coord, _truncate(raw, 100)))

        # 수식 셀
        if formula_cells:
            lines.append(f"### Formula cells ({len(formula_cells)}개)")
            lines.append("")
            lines.append("| 좌표 | 수식 | 계산값 (저장 당시) |")
            lines.append("|---|---|---|")
            for coord, formula, computed in formula_cells:
                lines.append(f"| {coord} | `{formula}` | {computed} |")
            lines.append("")
        else:
            lines.append(
                "### ⚠️ 수식 셀 없음 — dead xlsx 가능성"
            )
            lines.append("")
            lines.append(
                "산출 결과만 보존된 사본으로 보입니다. 원본(식 보존) 양식 별도 "
                "확보 필요. PDF/print용 사본이라면 운영자에게 LibreOffice·Excel "
                "원본 파일 요청."
            )
            lines.append("")

        # 값 셀 (수식 없는 텍스트/숫자 — 라벨·고정값·고정 옵션)
        if value_cells:
            lines.append(f"### Value cells ({len(value_cells)}개)")
            lines.append("")
            lines.append("| 좌표 | 값 |")
            lines.append("|---|---|")
            for coord, value in value_cells:
                lines.append(f"| {coord} | {value} |")
            lines.append("")

    out_text = "\n".join(lines)
    if out is not None:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(out_text, encoding="utf-8")
        print(f"[OK] dump 저장: {out}", file=sys.stderr)
    else:
        print(out_text)
    return out_text


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    xlsx_path = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx_path.exists():
        print(f"파일 없음: {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    if xlsx_path.suffix.lower() == ".xls":
        print(
            f"⚠️  {xlsx_path.suffix} 파일은 openpyxl이 못 읽음.\n"
            f"LibreOffice로 .xlsx 변환 후 다시 실행:\n"
            f"  soffice --headless --convert-to xlsx '{xlsx_path}'",
            file=sys.stderr,
        )
        sys.exit(2)
    out_path = Path(sys.argv[2]).expanduser() if len(sys.argv) >= 3 else None
    dump(xlsx_path, out_path)


if __name__ == "__main__":
    main()
